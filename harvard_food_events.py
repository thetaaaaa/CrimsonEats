"""
Harvard Food Events Scraper  v4
================================
抓取哈佛大学各院系活动页面中未来 7 天内提供餐饮的讲座/活动，输出为 Excel 表格。

覆盖来源（纯 HTML 解析，无 API 依赖）：
  - Harvard Law School        https://hls.harvard.edu/calendar/
  - Fairbank Center           https://fairbank.fas.harvard.edu/events/
  - IQSS                      https://www.iq.harvard.edu/calendar
  （可在 HTML_CALENDARS 中继续扩展）

工作原理：
  HLS        —— 自定义 WordPress 主题，直接在列表页显示完整摘要
  Fairbank   —— The Events Calendar 插件，餐饮信息在详情页全文中，需逐条进入详情页
  IQSS       —— Drupal 自定义主题，分页列表（?page=N），餐饮信息在详情页正文中

用法：
    pip install requests beautifulsoup4 lxml openpyxl
    python harvard_food_events.py

输出：
    harvard_food_events_YYYY-MM-DD.xlsx（与脚本同目录）
"""

import os
import re
import time
from datetime import datetime, timedelta, date
from zoneinfo import ZoneInfo

import requests
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════
# 全局配置
# ══════════════════════════════════════════════════════

BOSTON_TZ       = ZoneInfo("America/New_York")
DAYS_AHEAD      = 7        # 抓取未来几天的活动
REQUEST_TIMEOUT = 20       # 单次请求超时（秒）
SLEEP_BETWEEN   = 1.0      # 每次请求间隔（秒，礼貌性延迟）

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
}

# 餐饮关键词正则（匹配标题或全文）
FOOD_RE = re.compile(
    r"\b("
    r"food|lunch|luncheon|dinner|breakfast|brunch|supper|"
    r"refreshment|light refresh|catering|catered|"
    r"reception|cocktail|wine|beer|"
    r"snack|pizza|sandwich|buffet|potluck|bbq|barbecue|"
    r"meal|free food|"
    r"beverage|drink|coffee|tea"
    r")\b",
    re.IGNORECASE,
)

# 日历配置列表
HTML_CALENDARS = [
    {
        "name":    "Harvard Law School",
        "type":    "hls",
        "url":     "https://hls.harvard.edu/calendar/",
    },
    {
        "name":    "Fairbank Center for Chinese Studies",
        "type":    "tribe",          # The Events Calendar (WordPress plugin)
        "url":     "https://fairbank.fas.harvard.edu/events/",
        "detail":  True,             # 需要进入详情页才能看到完整餐饮信息
    },
    {
        "name":    "IQSS",
        "type":    "iqss",           # Drupal 自定义主题，分页 ?page=N
        "url":     "https://www.iq.harvard.edu/calendar",
        "base":    "https://www.iq.harvard.edu",
    },
    # ── 在此添加更多院系 ──
    # {"name": "Berkman Klein Center", "type": "tribe", "url": "https://cyber.harvard.edu/events", "detail": True},
]


# ══════════════════════════════════════════════════════
# 通用工具
# ══════════════════════════════════════════════════════

def _get(url: str, retries: int = 2) -> requests.Response | None:
    """带重试的 GET 请求"""
    for attempt in range(retries + 1):
        try:
            r = requests.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT)
            r.raise_for_status()
            return r
        except requests.exceptions.HTTPError as e:
            print(f"    [HTTP {e.response.status_code}] {url}")
            return None
        except requests.exceptions.ConnectionError:
            print(f"    [连接失败] {url}")
            return None
        except requests.exceptions.Timeout:
            if attempt < retries:
                print(f"    [超时，重试 {attempt+1}] {url}")
                time.sleep(2)
            else:
                print(f"    [超时放弃] {url}")
                return None
        except Exception as e:
            print(f"    [错误] {url}: {e}")
            return None
    return None


def _has_food(text: str) -> bool:
    return bool(FOOD_RE.search(text))


def _food_snippet(text: str, max_len: int = 140) -> str:
    """从全文中截取包含食物关键词的句子"""
    for sent in re.split(r'[.!?\n]', text):
        if FOOD_RE.search(sent):
            return sent.strip()[:max_len]
    return ""


def _parse_tribe_time(time_text: str, base_date: date) -> tuple:
    """
    解析 Tribe Events 时间文本，如 "April 6 @ 4:00 pm - 6:00 pm"
    返回 (start_datetime, end_datetime)
    """
    if not time_text:
        return None, None
    # 提取时间部分："4:00 pm - 6:00 pm" 或 "12:20 pm - 1:20 pm"
    m = re.search(r'@\s*(\d+:\d+\s*[ap]m)\s*[-–]\s*(\d+:\d+\s*[ap]m)', time_text, re.IGNORECASE)
    if not m:
        # 只有开始时间
        m2 = re.search(r'@\s*(\d+:\d+\s*[ap]m)', time_text, re.IGNORECASE)
        if not m2:
            return None, None
        try:
            start = datetime.strptime(m2.group(1).strip(), "%I:%M %p")
            start = datetime(base_date.year, base_date.month, base_date.day,
                             start.hour, start.minute, tzinfo=BOSTON_TZ)
            return start, None
        except Exception:
            return None, None
    try:
        s = datetime.strptime(m.group(1).strip(), "%I:%M %p")
        e = datetime.strptime(m.group(2).strip(), "%I:%M %p")
        start = datetime(base_date.year, base_date.month, base_date.day,
                         s.hour, s.minute, tzinfo=BOSTON_TZ)
        end   = datetime(base_date.year, base_date.month, base_date.day,
                         e.hour, e.minute, tzinfo=BOSTON_TZ)
        return start, end
    except Exception:
        return None, None


def _fmt(dt: datetime | None, fmt: str) -> str:
    return dt.strftime(fmt) if dt else ""


# ══════════════════════════════════════════════════════
# 爬虫 A：Harvard Law School（自定义 WP 主题）
# ══════════════════════════════════════════════════════
#
# 结构：
#   div.events-feed__group
#     p.events-feed__item-date        ← 日期 "April 3"
#     ul.events-feed__list
#       li.events-feed__item
#         h2.events-feed__item-title > a   ← 标题 + URL
#         p.events-feed__item-time         ← "12:00 pm - 1:30 pm"
#         div.events-feed__item-excerpt    ← 摘要（含餐饮信息）
#
# URL：/calendar/?start=YYYY-MM-DD  每次返回以该日为起点的一周

def fetch_hls(cal: dict, start_dt: datetime, end_dt: datetime) -> list[dict]:
    results, seen = [], set()
    current = start_dt.date()
    end     = end_dt.date()

    while current <= end:
        url = f"{cal['url']}?start={current.strftime('%Y-%m-%d')}"
        r   = _get(url)
        if not r:
            current += timedelta(days=7)
            continue

        soup = BeautifulSoup(r.text, "lxml")

        for group in soup.select("div.events-feed__group"):
            # 解析该组日期
            date_el  = group.select_one("p.events-feed__item-date")
            date_str = date_el.get_text(strip=True) if date_el else ""
            event_date = _parse_month_day(date_str, current.year)
            if event_date is None:
                continue
            if not (start_dt.date() <= event_date <= end):
                continue

            for li in group.select("li.events-feed__item"):
                a_tag = li.select_one("h2.events-feed__item-title a")
                if not a_tag:
                    continue

                title   = a_tag.get_text(" ", strip=True)
                ev_url  = a_tag.get("href", "")
                uid     = ev_url or title
                if uid in seen:
                    continue
                seen.add(uid)

                time_el  = li.select_one("p.events-feed__item-time")
                time_str = time_el.get_text(strip=True) if time_el else ""
                start_t, end_t = _parse_hls_time(time_str, event_date)

                desc_el = li.select_one("div.events-feed__item-excerpt")
                desc    = desc_el.get_text(" ", strip=True) if desc_el else ""
                full_text = title + " " + desc

                if not _has_food(full_text):
                    continue

                results.append({
                    "title":          title,
                    "start_datetime": start_t,
                    "end_datetime":   end_t,
                    "location":       "Harvard Law School",
                    "food_note":      _food_snippet(full_text),
                    "event_url":      ev_url,
                    "calendar":       cal["name"],
                })

        # HLS 每请求覆盖约 7 天
        current += timedelta(days=7)
        time.sleep(SLEEP_BETWEEN)

    return results


def _parse_hls_time(time_str: str, d: date) -> tuple:
    """解析 HLS 时间 '12:00 pm - 1:30 pm'"""
    if not time_str or not d:
        return None, None
    try:
        parts = [p.strip() for p in re.split(r'\s*[-–]\s*', time_str)]
        def to_dt(s):
            t = datetime.strptime(s, "%I:%M %p")
            return datetime(d.year, d.month, d.day, t.hour, t.minute, tzinfo=BOSTON_TZ)
        start = to_dt(parts[0]) if parts else None
        end   = to_dt(parts[1]) if len(parts) > 1 else None
        return start, end
    except Exception:
        fb = datetime(d.year, d.month, d.day, tzinfo=BOSTON_TZ)
        return fb, None


def _parse_month_day(s: str, year: int) -> date | None:
    """将 'April 3' 解析为 date"""
    for fmt in ("%B %d %Y", "%b %d %Y"):
        try:
            return datetime.strptime(f"{s} {year}", fmt).date()
        except ValueError:
            pass
    return None


# ══════════════════════════════════════════════════════
# 爬虫 B：The Events Calendar（Tribe Events）通用解析器
# ══════════════════════════════════════════════════════
#
# 适用站点（均使用 WordPress + The Events Calendar 插件）：
#   Fairbank Center   fairbank.fas.harvard.edu/events/
#
# 列表页结构：
#   article.tribe-events-calendar-list__event
#     .tribe-events-calendar-list__event-title > a     ← 标题 + URL
#     time[datetime]                                    ← date attr="YYYY-MM-DD"
#                                                         text="April 9 @ 12:20 pm - 1:20 pm"
#     .tribe-events-calendar-list__event-venue-title   ← 场馆名
#     .tribe-events-calendar-list__event-venue-address ← 地址
#     [class*="description"]                           ← 摘要（可能不含餐饮信息！）
#
# 详情页：餐饮信息在正文 <p> 中，需用 FOOD_RE 匹配。

def fetch_tribe(cal: dict, start_dt: datetime, end_dt: datetime) -> list[dict]:
    results, seen = [], set()
    needs_detail  = cal.get("detail", False)

    # 抓列表页（Tribe Events upcoming list 一次返回所有未来活动）
    r = _get(cal["url"])
    if not r:
        return results

    soup = BeautifulSoup(r.text, "lxml")

    for art in soup.select("article.tribe-events-calendar-list__event"):
        # 标题 & URL
        a_tag  = art.select_one(".tribe-events-calendar-list__event-title a")
        if not a_tag:
            continue
        title   = a_tag.get_text(" ", strip=True)
        ev_url  = a_tag.get("href", "")
        uid     = ev_url or title
        if uid in seen:
            continue
        seen.add(uid)

        # 日期过滤
        time_el   = art.select_one("time[datetime]")
        if not time_el:
            continue
        date_str  = time_el.get("datetime", "")   # "2026-04-09"
        try:
            event_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            continue
        if not (start_dt.date() <= event_date <= end_dt.date()):
            continue

        # 时间解析
        time_text = time_el.get_text(" ", strip=True)   # "April 9 @ 12:20 pm - 1:20 pm"
        start_t, end_t = _parse_tribe_time(time_text, event_date)

        # 地点
        venue_el = art.select_one(".tribe-events-calendar-list__event-venue-title")
        addr_el  = art.select_one(".tribe-events-calendar-list__event-venue-address")
        location = ""
        if venue_el:
            location = venue_el.get_text(strip=True)
        if addr_el:
            addr_short = addr_el.get_text(strip=True).split(",")[0]
            if addr_short and addr_short != location:
                location = f"{location}, {addr_short}".strip(", ")

        # 摘要（列表页）
        desc_el  = art.select_one("[class*='description']")
        list_desc = desc_el.get_text(" ", strip=True) if desc_el else ""

        # 判断是否需要进详情页
        food_note = ""
        if needs_detail and ev_url:
            # 先用列表摘要快速判断，若没命中再进详情页
            if _has_food(title + " " + list_desc):
                food_note = _food_snippet(title + " " + list_desc)
            else:
                food_note = _fetch_detail_food(ev_url)
                if not food_note:
                    continue   # 详情页也没有餐饮信息，跳过
        else:
            full_text = title + " " + list_desc
            if not _has_food(full_text):
                continue
            food_note = _food_snippet(full_text)

        results.append({
            "title":          title,
            "start_datetime": start_t,
            "end_datetime":   end_t,
            "location":       location or cal["name"],
            "food_note":      food_note,
            "event_url":      ev_url,
            "calendar":       cal["name"],
        })

        time.sleep(SLEEP_BETWEEN)

    return results


def _fetch_detail_food(url: str) -> str:
    """
    进入活动详情页，从全文中提取含餐饮关键词的句子。
    返回食物相关句子（非空即说明有餐饮），找不到则返回空字符串。
    """
    r = _get(url)
    if not r:
        return ""
    soup = BeautifulSoup(r.text, "lxml")

    # 取详情页主体文字（去掉导航、页脚等干扰区域）
    for tag in soup.select("nav, header, footer, .tribe-events-nav, script, style"):
        tag.decompose()

    body_text = soup.get_text(" ", strip=True)
    return _food_snippet(body_text)


# ══════════════════════════════════════════════════════
# 爬虫 C：IQSS（Drupal 自定义主题，分页列表）
# ══════════════════════════════════════════════════════
#
# 列表页（分页）：https://www.iq.harvard.edu/calendar?page=N
#   article[data-component-name="event-card"]
#     .event-card__heading              ← 标题
#     .event-card__link                 ← 相对 href，如 /event/slug
#     .event-card__date                 ← "Apr 8, 2026"
#     .event-card__time                 ← "12:00PM - 1:30PM EDT"
#
# 详情页：https://www.iq.harvard.edu/event/slug
#   .event-details__description         ← 含餐饮信息的正文
#   .field--name-field-hwp-event-location ← 地点
#
# 分页：?page=0 起，遇到所有活动均超出日期范围时停止

def fetch_iqss(cal: dict, start_dt: datetime, end_dt: datetime) -> list[dict]:
    base    = cal.get("base", "https://www.iq.harvard.edu")
    results, seen = [], set()

    page = 0
    while True:
        url = f"{cal['url']}?page={page}"
        r   = _get(url)
        if not r:
            break

        soup  = BeautifulSoup(r.text, "lxml")
        cards = soup.select('article[data-component-name="event-card"]')
        if not cards:
            break

        page_has_valid = False   # 该页是否有在日期范围内的活动
        all_past       = True    # 该页所有活动是否都已过期（用于提前终止）

        for card in cards:
            heading  = card.select_one(".event-card__heading")
            link_el  = card.select_one(".event-card__link")
            date_el  = card.select_one(".event-card__date")
            time_el  = card.select_one(".event-card__time")

            if not heading or not date_el:
                continue

            title    = heading.get_text(strip=True)
            href     = link_el.get("href", "") if link_el else ""
            # href 可能是相对路径或已含 base
            if href.startswith("/"):
                ev_url = base + href.split("?")[0]   # 去掉 ?occ_id=0 等参数
            elif href.startswith("http"):
                ev_url = href.split("?")[0]
            else:
                ev_url = ""

            # 日期解析："Apr 8, 2026"
            date_str   = date_el.get_text(strip=True)
            event_date = _parse_iqss_date(date_str)
            if event_date is None:
                continue

            if event_date > end_dt.date():
                continue          # 超出范围，跳过（但不能提前终止，因为列表不严格按日期排序）
            if event_date < start_dt.date():
                continue          # 已过期
            all_past = False

            # 时间解析："12:00PM - 1:30PM EDT"
            time_str         = time_el.get_text(strip=True) if time_el else ""
            start_t, end_t   = _parse_iqss_time(time_str, event_date)

            uid = ev_url or title
            if uid in seen:
                continue
            seen.add(uid)
            page_has_valid = True

            # 进详情页获取地点和餐饮信息
            location  = "IQSS"
            food_note = ""
            if ev_url:
                location, food_note = _fetch_iqss_detail(ev_url)

            if not food_note:
                continue          # 无餐饮信息，跳过

            results.append({
                "title":          title,
                "start_datetime": start_t,
                "end_datetime":   end_t,
                "location":       location,
                "food_note":      food_note,
                "event_url":      ev_url,
                "calendar":       cal["name"],
            })
            time.sleep(SLEEP_BETWEEN)

        # 翻页控制：若本页所有活动均未来且超出范围，停止
        if not cards or (not page_has_valid and all_past):
            break

        # 检查是否有下一页
        next_link = soup.select_one('.hwp-pager a[rel="prev"]')   # 注意：IQSS 的 rel 标注反了
        if not next_link:
            # 备用：找页码中最大的 page= 参数
            pager_links = soup.select('.hwp-pager a')
            max_page = page
            for a in pager_links:
                href_p = a.get("href", "")
                m = re.search(r"page=(\d+)", href_p)
                if m:
                    max_page = max(max_page, int(m.group(1)))
            if max_page <= page:
                break
        page += 1
        time.sleep(SLEEP_BETWEEN)

    return results


def _parse_iqss_date(s: str) -> date | None:
    """解析 'Apr 8, 2026' → date"""
    s = s.strip()
    for fmt in ("%b %d, %Y", "%B %d, %Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_iqss_time(time_str: str, d: date) -> tuple:
    """
    解析 '12:00PM - 1:30PM EDT' → (start_datetime, end_datetime)
    """
    if not time_str or not d:
        return None, None
    # 去掉时区后缀
    clean = re.sub(r'\s+[A-Z]{2,4}$', '', time_str.strip())
    parts = [p.strip() for p in re.split(r'\s*[-–]\s*', clean)]
    def to_dt(s):
        for fmt in ("%I:%M%p", "%I:%M %p"):
            try:
                t = datetime.strptime(s.strip(), fmt)
                return datetime(d.year, d.month, d.day, t.hour, t.minute,
                                tzinfo=BOSTON_TZ)
            except ValueError:
                pass
        return None
    start = to_dt(parts[0]) if len(parts) >= 1 else None
    end   = to_dt(parts[1]) if len(parts) >= 2 else None
    return start, end


def _fetch_iqss_detail(url: str) -> tuple[str, str]:
    """
    抓取 IQSS 活动详情页，返回 (location, food_note)。
    location：.field--name-field-hwp-event-location 的文本
    food_note：含餐饮关键词的句子
    """
    r = _get(url)
    if not r:
        return "IQSS", ""

    soup = BeautifulSoup(r.text, "lxml")

    # 地点
    loc_el   = soup.select_one(".field--name-field-hwp-event-location")
    location = loc_el.get_text(" ", strip=True) if loc_el else "IQSS"
    # 去掉 "pin_drop Location" 等 Material Icon 文字
    location = re.sub(r'^(pin_drop|location_on|place)\s*', '', location, flags=re.IGNORECASE).strip()
    location = re.sub(r'^Location\s*', '', location, flags=re.IGNORECASE).strip()

    # IQSS 详情页的文字分布在多个字段中：
    #   .field--name-field-hwp-introduction  ← 活动介绍（含餐饮信息）
    #   .event-details__description          ← 演讲摘要（Speaker/Abstract）
    #   .field--name-field-hwp-body          ← 备用正文
    # 将所有字段文本合并后匹配
    field_sels = [
        ".field--name-field-hwp-introduction",
        ".event-details__description",
        ".field--name-field-hwp-body",
    ]
    body_parts = []
    for sel in field_sels:
        el = soup.select_one(sel)
        if el:
            body_parts.append(el.get_text(" ", strip=True))
    body_text = " ".join(body_parts)

    # 若上述字段都不含食物词，退一步搜全文
    if not _has_food(body_text):
        for tag in soup.select("nav, header, footer, .hwp-pager, script, style"):
            tag.decompose()
        body_text = soup.get_text(" ", strip=True)

    food_note = _food_snippet(body_text)
    return location, food_note


# ══════════════════════════════════════════════════════
# 分发器
# ══════════════════════════════════════════════════════

SCRAPER_MAP = {
    "hls":   fetch_hls,
    "tribe": fetch_tribe,
    "iqss":  fetch_iqss,
}

def run_all(start_dt: datetime, end_dt: datetime) -> list[dict]:
    all_events, seen_urls = [], set()

    for cal in HTML_CALENDARS:
        scraper = SCRAPER_MAP.get(cal["type"])
        if not scraper:
            print(f"  [!] 未知 scraper 类型: {cal['type']}")
            continue

        print(f"\n🔍 抓取: {cal['name']} ...")
        try:
            events = scraper(cal, start_dt, end_dt)
        except Exception as e:
            print(f"  [!] 抓取出错: {e}")
            events = []

        new_count = 0
        for ev in events:
            uid = ev.get("event_url") or ev.get("title", "")
            if uid not in seen_urls:
                seen_urls.add(uid)
                all_events.append(ev)
                new_count += 1
        print(f"   ✓ 新增 {new_count} 个有餐饮的活动")

    return all_events


# ══════════════════════════════════════════════════════
# Excel 输出
# ══════════════════════════════════════════════════════

def write_excel(events: list[dict], path: str, start_dt: datetime):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Food Events"

    # ── 样式 ──
    DEEP_BLUE  = PatternFill("solid", fgColor="1B3A6B")
    CRIMSON    = PatternFill("solid", fgColor="A51C30")
    STRIPE_A   = PatternFill("solid", fgColor="EEF2FB")
    STRIPE_B   = PatternFill("solid", fgColor="FFFFFF")
    HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name="Calibri", bold=True, color="1B3A6B", size=14)
    SUB_FONT   = Font(name="Calibri", italic=True, color="666666", size=10)
    BODY_FONT  = Font(name="Calibri", size=10)
    LINK_FONT  = Font(name="Calibri", size=10, color="1155CC", underline="single")
    NOTE_FONT  = Font(name="Calibri", size=10, color="A51C30")
    thin = Side(style="thin", color="CCCCCC")
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── 标题行 ──
    COL_COUNT = 8
    ws.merge_cells(f"A1:{get_column_letter(COL_COUNT)}1")
    ws["A1"] = "🎓  Harvard University — Upcoming Events with Food & Refreshments"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{get_column_letter(COL_COUNT)}2")
    end_dt = start_dt + timedelta(days=DAYS_AHEAD)
    ws["A2"] = (
        f"生成时间: {datetime.now(BOSTON_TZ).strftime('%Y-%m-%d  %H:%M ET')}　｜　"
        f"查询范围: {start_dt.strftime('%m/%d')} – {end_dt.strftime('%m/%d/%Y')}　｜　"
        f"共 {len(events)} 个活动"
    )
    ws["A2"].font      = SUB_FONT
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── 列标题 ──
    COLS = [
        ("活动名称",   44),
        ("日期",       13),
        ("开始",       9),
        ("结束",       9),
        ("地点",       28),
        ("餐饮信息",   34),
        ("来源",       24),
        ("链接",       10),
    ]
    for ci, (hdr, w) in enumerate(COLS, 1):
        c = ws.cell(row=3, column=ci, value=hdr)
        c.font      = HDR_FONT
        c.fill      = DEEP_BLUE
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[3].height = 22

    # ── 数据行 ──
    sorted_evs = sorted(
        events,
        key=lambda e: e.get("start_datetime") or datetime(1970, 1, 1, tzinfo=BOSTON_TZ)
    )

    for ri, ev in enumerate(sorted_evs, start=4):
        fill = STRIPE_A if ri % 2 == 0 else STRIPE_B
        s    = ev.get("start_datetime")
        e    = ev.get("end_datetime")

        row = [
            ev.get("title", ""),
            _fmt(s, "%m/%d (%a)"),
            _fmt(s, "%H:%M") if s else "",
            _fmt(e, "%H:%M") if e else "",
            ev.get("location", ""),
            ev.get("food_note", ""),
            ev.get("calendar", ""),
            "",
        ]
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = fill
            c.border    = bdr
            c.alignment = Alignment(vertical="top", wrap_text=(ci in (1, 5, 6)))
            # 餐饮信息列用红色字体突出
            c.font = NOTE_FONT if ci == 6 else BODY_FONT

        # 链接列
        url = ev.get("event_url", "")
        lc  = ws.cell(row=ri, column=8)
        if url:
            lc.value     = "详情 →"
            lc.hyperlink = url
            lc.font      = LINK_FONT
        lc.fill      = fill
        lc.alignment = Alignment(horizontal="center", vertical="top")
        lc.border    = bdr
        ws.row_dimensions[ri].height = 40

    ws.freeze_panes    = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(COL_COUNT)}{3 + len(sorted_evs)}"

    # ── 说明 Sheet ──
    ws2 = wb.create_sheet("说明")
    notes = [
        ("脚本版本",   "v3  (纯 HTML 解析，无 API)"),
        ("生成时间",   datetime.now(BOSTON_TZ).strftime("%Y-%m-%d %H:%M ET")),
        ("查询范围",   f"未来 {DAYS_AHEAD} 天"),
        ("数据来源",   "\n".join(c["name"] for c in HTML_CALENDARS)),
        ("判断逻辑",   "HLS: 列表页摘要匹配关键词\n"
                      "Fairbank: 进入每个活动详情页匹配全文\n"
                      "IQSS: 分页列表(?page=N)+进入详情页匹配.event-details__description"),
        ("餐饮关键词", "lunch / dinner / breakfast / brunch / reception / refreshment\n"
                      "snack / pizza / coffee / tea / wine / beer / beverage / meal …"),
        ("运行方式",   "pip install requests beautifulsoup4 lxml openpyxl\n"
                      "python harvard_food_events.py"),
        ("注意事项",   "• 需要哈佛校园网或 VPN（或有公开访问权限的网络）\n"
                      "• 请遵守各网站服务条款，勿频繁请求\n"
                      "• 建议每天运行一次以获取最新信息"),
        ("如需扩展",   "在脚本顶部 HTML_CALENDARS 列表中添加新条目即可"),
    ]
    for ri, (k, v) in enumerate(notes, 1):
        kc = ws2.cell(ri, 1, k)
        kc.font = Font(bold=True, name="Calibri", size=10)
        kc.fill = PatternFill("solid", fgColor="F0F0F0")
        vc = ws2.cell(ri, 2, v)
        vc.font      = Font(name="Calibri", size=10)
        vc.alignment = Alignment(wrap_text=True)
        ws2.row_dimensions[ri].height = max(18, v.count("\n") * 16 + 18)
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 68

    wb.save(path)
    print(f"\n✅  Excel 已保存：{path}")


# ══════════════════════════════════════════════════════
# 主程序
# ══════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════
# HTML 输出（供 GitHub Pages 使用）
# ══════════════════════════════════════════════════════

def write_html(events: list[dict], path: str, now: datetime) -> None:
    """
    生成一个自包含的 HTML 页面，展示有餐饮的活动列表。
    固定文件名（food-events.html），每次覆盖，GitHub Pages 直接服务。
    """
    end_dt     = now + timedelta(days=DAYS_AHEAD)
    updated_et = now.strftime("%Y-%m-%d %H:%M ET")
    sorted_evs = sorted(
        events,
        key=lambda e: e.get("start_datetime") or datetime(1970, 1, 1, tzinfo=BOSTON_TZ)
    )

    def esc(s: str) -> str:
        return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")

    # 按日期分组
    from collections import defaultdict
    by_date: dict[str, list] = defaultdict(list)
    no_date = []
    for ev in sorted_evs:
        s = ev.get("start_datetime")
        key = s.strftime("%Y-%m-%d") if s else ""
        if key:
            by_date[key].append(ev)
        else:
            no_date.append(ev)

    rows_html = ""
    for date_key in sorted(by_date.keys()):
        evs_on_day = by_date[date_key]
        dt_obj     = datetime.strptime(date_key, "%Y-%m-%d").replace(tzinfo=BOSTON_TZ)
        date_label = dt_obj.strftime("%A, %B %-d")   # e.g. "Wednesday, April 8"
        rows_html += f'<tr class="date-row"><td colspan="5">{esc(date_label)}</td></tr>\n'
        for ev in evs_on_day:
            s = ev.get("start_datetime")
            e = ev.get("end_datetime")
            time_str = ""
            if s:
                time_str = s.strftime("%-I:%M %p")
                if e:
                    time_str += f" – {e.strftime('%-I:%M %p')}"
            url   = esc(ev.get("event_url", ""))
            title = esc(ev.get("title", ""))
            title_cell = f'<a href="{url}" target="_blank" rel="noopener">{title}</a>' if url else title
            rows_html += (
                f'<tr>'
                f'<td class="time">{esc(time_str)}</td>'
                f'<td class="title">{title_cell}</td>'
                f'<td class="food">{esc(ev.get("food_note", ""))}</td>'
                f'<td class="loc">{esc(ev.get("location", ""))}</td>'
                f'<td class="src">{esc(ev.get("calendar", ""))}</td>'
                f'</tr>\n'
            )
    if no_date:
        rows_html += '<tr class="date-row"><td colspan="5">Date TBD</td></tr>\n'
        for ev in no_date:
            url   = esc(ev.get("event_url", ""))
            title = esc(ev.get("title", ""))
            title_cell = f'<a href="{url}" target="_blank" rel="noopener">{title}</a>' if url else title
            rows_html += (
                f'<tr>'
                f'<td class="time">—</td>'
                f'<td class="title">{title_cell}</td>'
                f'<td class="food">{esc(ev.get("food_note", ""))}</td>'
                f'<td class="loc">{esc(ev.get("location", ""))}</td>'
                f'<td class="src">{esc(ev.get("calendar", ""))}</td>'
                f'</tr>\n'
            )

    empty_msg = ""
    if not events:
        empty_msg = '<p class="empty">No food events found for the next 7 days. Check back tomorrow!</p>'

    sources_list = "".join(
        f'<li><a href="{c["url"]}" target="_blank">{esc(c["name"])}</a></li>'
        for c in HTML_CALENDARS
    )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Free Food at Harvard — Next 7 Days</title>
  <style>
    :root {{
      --crimson: #A51C30;
      --navy:    #1B3A6B;
      --bg:      #F9F8F6;
      --card:    #FFFFFF;
      --border:  #E0DDD8;
      --text:    #1a1a1a;
      --muted:   #666;
      --food:    #B7350F;
    }}
    * {{ box-sizing: border-box; margin: 0; padding: 0; }}
    body {{
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
      background: var(--bg);
      color: var(--text);
      padding: 2rem 1rem;
    }}
    header {{
      max-width: 900px;
      margin: 0 auto 2rem;
      border-left: 4px solid var(--crimson);
      padding-left: 1rem;
    }}
    header h1 {{
      font-size: 1.6rem;
      color: var(--navy);
      margin-bottom: .3rem;
    }}
    header h1 span {{ color: var(--crimson); }}
    .meta {{ font-size: .82rem; color: var(--muted); }}
    .meta a {{ color: var(--crimson); text-decoration: none; }}
    .container {{
      max-width: 900px;
      margin: 0 auto;
      background: var(--card);
      border: 1px solid var(--border);
      border-radius: 8px;
      overflow: hidden;
      box-shadow: 0 2px 8px rgba(0,0,0,.06);
    }}
    table {{ width: 100%; border-collapse: collapse; font-size: .9rem; }}
    th {{
      background: var(--navy);
      color: #fff;
      font-weight: 600;
      text-align: left;
      padding: .65rem .9rem;
      font-size: .8rem;
      letter-spacing: .04em;
      text-transform: uppercase;
    }}
    tr.date-row td {{
      background: #EEF2FB;
      color: var(--navy);
      font-weight: 700;
      padding: .5rem .9rem;
      font-size: .85rem;
      border-top: 2px solid var(--border);
    }}
    td {{
      padding: .6rem .9rem;
      border-bottom: 1px solid var(--border);
      vertical-align: top;
    }}
    tr:last-child td {{ border-bottom: none; }}
    tr:not(.date-row):hover td {{ background: #fafaf8; }}
    td.time  {{ white-space: nowrap; color: var(--muted); font-size: .82rem; width: 110px; }}
    td.title a {{ color: var(--navy); text-decoration: none; font-weight: 500; }}
    td.title a:hover {{ text-decoration: underline; color: var(--crimson); }}
    td.food  {{ color: var(--food); font-size: .82rem; font-style: italic; }}
    td.loc   {{ color: var(--muted); font-size: .8rem; }}
    td.src   {{
      color: var(--muted); font-size: .78rem;
      white-space: nowrap;
    }}
    .empty {{ padding: 2rem; text-align: center; color: var(--muted); font-style: italic; }}
    footer {{
      max-width: 900px;
      margin: 1.5rem auto 0;
      font-size: .78rem;
      color: var(--muted);
    }}
    footer ul {{ list-style: none; display: flex; gap: 1rem; flex-wrap: wrap; margin-top: .3rem; }}
    footer a {{ color: var(--crimson); text-decoration: none; }}
    @media (max-width: 600px) {{
      td.loc, td.src {{ display: none; }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>🎓 Free Food at Harvard — <span>Next 7 Days</span></h1>
    <p class="meta">
      {now.strftime("%B %-d")} – {end_dt.strftime("%B %-d, %Y")} &nbsp;·&nbsp;
      Auto-updated daily &nbsp;·&nbsp;
      Last updated: <strong>{updated_et}</strong>
    </p>
  </header>

  <div class="container">
    {empty_msg}
    {'<table><thead><tr><th>Time</th><th>Event</th><th>Food</th><th>Location</th><th>Source</th></tr></thead><tbody>' + rows_html + '</tbody></table>' if events else ''}
  </div>

  <footer>
    <p>Sources monitored:</p>
    <ul>{sources_list}</ul>
  </footer>
</body>
</html>
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"✅  HTML 已保存：{path}")


# ══════════════════════════════════════════════════════
# README 输出（供 GitHub 仓库首页展示）
# ══════════════════════════════════════════════════════

def write_readme(events: list[dict], path: str, now: datetime) -> None:
    """
    将有餐饮的活动列表写入 README.md，按日期分组展示为 Markdown 表格。
    固定文件名，每次覆盖，提交后 GitHub 仓库首页直接展示。
    """
    from collections import defaultdict

    end_dt     = now + timedelta(days=DAYS_AHEAD)
    updated_et = now.strftime("%Y-%m-%d %H:%M ET")

    sorted_evs = sorted(
        events,
        key=lambda e: e.get("start_datetime") or datetime(1970, 1, 1, tzinfo=BOSTON_TZ)
    )

    # 按日期分组
    by_date: dict[str, list] = defaultdict(list)
    no_date = []
    for ev in sorted_evs:
        s   = ev.get("start_datetime")
        key = s.strftime("%Y-%m-%d") if s else ""
        if key:
            by_date[key].append(ev)
        else:
            no_date.append(ev)

    lines = []
    lines.append("# 🎓 Free Food at Harvard — Next 7 Days")
    lines.append("")
    lines.append(
        f"> **{now.strftime('%B %-d')} – {end_dt.strftime('%B %-d, %Y')}** &nbsp;·&nbsp; "
        f"Auto-updated daily via GitHub Actions &nbsp;·&nbsp; "
        f"Last updated: **{updated_et}**"
    )
    lines.append("")

    if not events:
        lines.append("*No food events found for the next 7 days. Check back tomorrow!*")
    else:
        def md_esc(s: str) -> str:
            """转义 Markdown 表格中的竖线"""
            return (s or "").replace("|", "\\|")

        for date_key in sorted(by_date.keys()):
            evs_on_day = by_date[date_key]
            dt_obj     = datetime.strptime(date_key, "%Y-%m-%d").replace(tzinfo=BOSTON_TZ)
            date_label = dt_obj.strftime("%A, %B %-d")

            lines.append(f"## {date_label}")
            lines.append("")
            lines.append("| Time | Event | Food | Location | Source |")
            lines.append("|------|-------|------|----------|--------|")

            for ev in evs_on_day:
                s = ev.get("start_datetime")
                e = ev.get("end_datetime")
                if s:
                    time_str = s.strftime("%-I:%M %p")
                    if e:
                        time_str += f" – {e.strftime('%-I:%M %p')}"
                else:
                    time_str = "—"

                url   = ev.get("event_url", "")
                title = md_esc(ev.get("title", ""))
                title_cell = f"[{title}]({url})" if url else title

                food     = md_esc(ev.get("food_note", ""))
                location = md_esc(ev.get("location", ""))
                source   = md_esc(ev.get("calendar", ""))

                lines.append(f"| {time_str} | {title_cell} | {food} | {location} | {source} |")

            lines.append("")

        if no_date:
            lines.append("## Date TBD")
            lines.append("")
            lines.append("| Event | Food | Location | Source |")
            lines.append("|-------|------|----------|--------|")
            for ev in no_date:
                url   = ev.get("event_url", "")
                title = md_esc(ev.get("title", ""))
                title_cell = f"[{title}]({url})" if url else title
                food     = md_esc(ev.get("food_note", ""))
                location = md_esc(ev.get("location", ""))
                source   = md_esc(ev.get("calendar", ""))
                lines.append(f"| {title_cell} | {food} | {location} | {source} |")
            lines.append("")

    # 数据来源
    lines.append("---")
    lines.append("")
    lines.append("**Sources monitored:**")
    for c in HTML_CALENDARS:
        lines.append(f"- [{c['name']}]({c['url']})")
    lines.append("")
    lines.append(
        "*This page is generated automatically by "
        "[`harvard_food_events.py`](harvard_food_events.py). "
        "Run the script locally or let GitHub Actions update it daily.*"
    )
    
    # 日历订阅部分（保留）
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("### 📅 Subscribe to Calendar")
    lines.append("")
    lines.append("```")
    lines.append("https://raw.githubusercontent.com/thetaaaaa/CrimsonEats/main/events.ics")
    lines.append("```")
    
    # 免责声明部分（保留）
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## ⚖️ Disclaimer")
    lines.append("")
    lines.append("This project aggregates publicly available event information from official Harvard University calendars for student convenience. All content is sourced from and remains the intellectual property of Harvard University and its respective schools/centers.")
    lines.append("")
    lines.append("**Non-commercial use only.** This is an independent student project, not officially affiliated with Harvard University. Always verify event details on official Harvard websites before attending.")
    lines.append("")
    lines.append("**Liability:** Information is provided \"as-is\" without warranties. Events may change, be cancelled, or have restricted attendance. Project maintainers are not responsible for outdated information or event-related issues.")
    lines.append("")
    lines.append("If Harvard University requests changes to this project, such requests will be honored promptly.")

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")
    print(f"✅  README 已保存：{path}")


# ══════════════════════════════════════════════════════
# 主程序
# ══════════════════════════════════════════════════════

def main():
    now    = datetime.now(BOSTON_TZ)
    end_dt = now + timedelta(days=DAYS_AHEAD)

    print("=" * 64)
    print("🎓  Harvard Food Events Scraper  v4")
    print(f"    日期范围: {now.strftime('%Y-%m-%d')} → {end_dt.strftime('%Y-%m-%d')}")
    print(f"    来源数量: {len(HTML_CALENDARS)} 个")
    print("=" * 64)

    events = run_all(now, end_dt)

    print(f"\n{'='*64}")
    print(f"    共找到 {len(events)} 个有餐饮的活动")

    # 输出路径与脚本同目录
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # Excel（本地用）
    xlsx_path = os.path.join(
        script_dir, f"harvard_food_events_{now.strftime('%Y-%m-%d')}.xlsx"
    )
    write_excel(events, xlsx_path, now)

    # README（GitHub 仓库首页，固定文件名，每次覆盖）
    readme_path = os.path.join(script_dir, "README.md")
    write_readme(events, readme_path, now)

    # 终端摘要
    if events:
        sorted_evs = sorted(
            events,
            key=lambda e: e.get("start_datetime") or datetime(1970, 1, 1, tzinfo=BOSTON_TZ)
        )
        print(f"\n{'日期':<12} {'时间':<7} {'名称':<42} {'餐饮'}")
        print("─" * 85)
        for ev in sorted_evs:
            s = ev.get("start_datetime")
            print(
                f"{_fmt(s,'%m/%d(%a)'):<12} "
                f"{_fmt(s,'%H:%M'):<7} "
                f"{ev['title'][:40]:<42} "
                f"{ev['food_note'][:30]}"
            )
    else:
        print("\n⚠️   未找到结果。")


if __name__ == "__main__":
    main()
